import http.client
from bs4 import BeautifulSoup
import pandas as pd
import mysql.connector
from openpyxl import load_workbook


# === 1. 发送请求获取HTML内容 ===
def get_page_content(page_num):
    conn = http.client.HTTPSConnection('detail.zol.com.cn')
    headers = {
        'accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7',
        'accept-language': 'zh-CN,zh;q=0.9,en;q=0.8,en-GB;q=0.7,en-US;q=0.6',
        'user-agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/137.0.0.0 Safari/537.36 Edg/137.0.0.0',
        'cookie': 'ip_ck=78aA4PP+j7QuODExNzk5LjE3NDIwOTIzMzE%3D; Hm_lvt_84a68916f17b61bfae5954e308133587=1742092332; Adshow=5; lv=1749093636; vn=4; Hm_lvt_ae5edc2bc4fc71370807f6187f0a2dd0=1747828481,1747970695,1749093637; Hm_lpvt_ae5edc2bc4fc71370807f0a2dd0=1749093637; HMACCOUNT=EE288EA3C07F1C87; z_pro_city=s_provice%3Dshanghai%26s_city%3Dshanghai; z_day=ixgo20=1&rdetail=1; realLocationId=26; userFidLocationId=26; questionnaire_pv=1749081602; listSubcateId=28',
    }
    path = f'/cpu/{page_num}.html' if page_num > 1 else '/cpu/'
    conn.request('GET', path, headers=headers)
    response = conn.getresponse()
    content = response.read().decode('gbk', errors='ignore')
    conn.close()
    print(f"成功获取第{page_num}页，内容长度: {len(content)}")
    return content


# === 2. 解析HTML提取CPU数据 ===
def parse_cpu_data(content):
    soup = BeautifulSoup(content, 'html.parser')
    cpu_data = []
    product_list = soup.find('ul', id='J_PicMode')

    if not product_list:
        print("警告：未找到商品列表，可能页面结构已变化")
        return cpu_data

    for li in product_list.find_all('li', class_=lambda x: x and 'nth-child' in x):
        item = {
            '名称': 'N/A',
            '参考价': 'N/A',
            '评分': '0',
            '点评数量': '0',
            '京东价格': 'N/A',
            '天猫价格': 'N/A',
            '详情页URL': '#'
        }

        # 提取名称
        h3 = li.find('h3')
        if h3:
            item['名称'] = h3.text.strip().split('<span>')[0].strip()

        # 修复参考价提取逻辑
        price_span = li.find('span', class_='price price-normal')
        if price_span:
            price_type = price_span.find('b', class_='price-type')
            if price_type:
                item['参考价'] = price_type.text.strip()
            else:
                # 尝试直接获取价格文本
                price_text = price_span.get_text(strip=True).replace('￥', '')
                if price_text:
                    item['参考价'] = price_text

        # 提取评分和点评数量
        comment_row = li.find('div', class_='comment-row')
        if comment_row:
            score_span = comment_row.find('span', class_='score')
            item['评分'] = score_span.text.strip() if score_span else '0'
            comment_link = comment_row.find('a', class_='comment-num')
            item['点评数量'] = comment_link.text.strip().replace('人点评', '') if comment_link else '0'

        # 提取电商价格
        b2c_price = li.find('div', class_='item-b2cprice')
        if b2c_price:
            jd_a = b2c_price.find('a', href=lambda x: 'jd.com' in x)
            item['京东价格'] = jd_a.text.strip() if jd_a else 'N/A'
            tmall_a = b2c_price.find('a', href=lambda x: 'taobao.com' in x or 'tmall.com' in x)
            item['天猫价格'] = tmall_a.text.strip() if tmall_a else 'N/A'

        # 提取详情页URL
        detail_a = li.find('a', class_='pic', href=True)
        if detail_a:
            item['详情页URL'] = f'https://detail.zol.com.cn{detail_a["href"]}' if detail_a["href"].startswith('/') else \
            detail_a["href"]

        cpu_data.append(item)

    print(f"当前页解析到 {len(cpu_data)} 条CPU数据")
    return cpu_data


# === 3. 主程序：循环爬取多页数据 ===
start_page = 1  # 起始页码
end_page = 65  # 结束页码（可根据需要修改）

all_cpu_data = []

for page_num in range(start_page, end_page + 1):
    content = get_page_content(page_num)
    page_data = parse_cpu_data(content)
    all_cpu_data.extend(page_data)

print(f"\n总共解析到 {len(all_cpu_data)} 条CPU数据（来自 {end_page - start_page + 1} 页）")

# === 4. 保存至Excel ===
columns_order = ['名称', '参考价', '评分', '点评数量', '京东价格', '天猫价格', '详情页URL']
df = pd.DataFrame(all_cpu_data, columns=columns_order)
df.to_excel('cpu_prices_all.xlsx', index=False)
print(f"Excel文件已生成，包含 {len(df)} 行数据")

# === 5. 导入MySQL ===
try:
    conn = mysql.connector.connect(
        host='localhost',
        user='root',
        password='123456',
        database='cpu_info'
    )
    cursor = conn.cursor()
    print("成功连接到MySQL")

    # 创建表（移除页码字段）
    cursor.execute('''
    CREATE TABLE IF NOT EXISTS cpu_prices (
        id INT AUTO_INCREMENT PRIMARY KEY,
        名称 VARCHAR(255),
        参考价 VARCHAR(50),
        评分 VARCHAR(10),
        点评数量 VARCHAR(50),
        京东价格 VARCHAR(50),
        天猫价格 VARCHAR(50),
        详情页URL TEXT
    )
    ''')
    print("表已创建或已存在")

    wb = load_workbook('cpu_prices_all.xlsx')
    ws = wb.active
    print("Excel文件读取成功，开始插入数据")

    for row in ws.iter_rows(min_row=2, values_only=True):  # 跳过标题行
        # 确保数据完整性
        if len(row) < 7:
            # 如果数据不足7列，补足到7列
            row = list(row) + ['N/A'] * (7 - len(row))
        
        # 显式指定每个字段的值
        name = row[0] if row[0] else 'N/A'
        price = row[1] if row[1] else 'N/A'
        score = row[2] if row[2] else '0'
        comment_count = row[3] if row[3] else '0'
        jd_price = row[4] if row[4] else 'N/A'
        tmall_price = row[5] if row[5] else 'N/A'
        url = row[6] if row[6] else '#'
        
        cursor.execute('''
        INSERT INTO cpu_prices (名称, 参考价, 评分, 点评数量, 京东价格, 天猫价格, 详情页URL)
        VALUES (%s, %s, %s, %s, %s, %s, %s)
        ''', (name, price, score, comment_count, jd_price, tmall_price, url))

    conn.commit()
    cursor.close()
    conn.close()
    print("数据成功导入MySQL")

except Exception as e:
    print(f"MySQL导入失败: {e}")
    # 调试：打印具体行数据和参数数量
    if 'row' in locals() and row:
        print(f"错误行数据: {row}, 参数数量: {len(row)}")